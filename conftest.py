import re
import pandas as pd
from datetime import datetime
from collections import OrderedDict
import pytest
from _pytest.mark.structures import Mark
import openpyxl
from openpyxl.drawing.image import Image
_py_ext_re = re.compile(r"\.py$")


def pytest_addoption(parser):
    group = parser.getgroup("terminal reporting")
    group.addoption('--reportexcel', '--report-excel',
                    action="store",
                    dest="excelpath",
                    metavar="path",
                    default=None,
                    help="create excel report file at given path.")
    
    parser.addoption('--executor', '--test-executor',
                    action="store",
                    dest="executor",
                    metavar="executor",
                    default=None,
                    help="create name of executor in excel file.")
    
    parser.addoption('--function', '--test-function',
                    action="store",
                    dest="function",
                    metavar="function",
                    default=None,
                    help="create name of function need to test in excel file.")
    
    parser.addoption('--logo', '--test-logo',
                    action="store",
                    dest="logo",
                    metavar="logo",
                    default=None,
                    help="create a logo for test file in excel file.")


def pytest_configure(config):
    excelpath = config.option.excelpath
    executor = config.option.executor
    function = config.option.function
    logo = config.option.logo
    if excelpath:
        config._excel = ExcelReporterV2(excelpath, executor, function, logo)
        config.pluginmanager.register(config._excel)


def pytest_unconfigure(config):
    excel = getattr(config, '_excel', None)
    if excel:
        del config._excel
        config.pluginmanager.unregister(excel)


def mangle_test_address(address):
    path, possible_open_bracket, params = address.partition('[')
    names = path.split("::")
    try:
        names.remove('()')
    except ValueError:
        pass

    names[0] = names[0].replace("/", '.')
    names[0] = _py_ext_re.sub("", names[0])
    names[-1] += possible_open_bracket + params
    return names


class ExcelReporterV2(object):


    def __init__(self, excelpath, executor, function, logo):
        self.results = []
        self.failed_count = 0
        self.success_count = 0
        self.skip_count = 0
        self.excelpath = datetime.now().strftime(excelpath)
        setattr(self, "executor", executor)
        setattr(self, "function", function)
        setattr(self, "logo", logo)

    def append(self, result):
        self.results.append(result)


    def create_sheet(self, column_heading):
        self.wbook = pd.DataFrame(columns = column_heading)

    def update_worksheet(self):
        self.wbook = pd.concat([self.wbook, pd.DataFrame(self.results)], ignore_index = False)

    def save_excel(self):
        self.wbook.to_excel(self.excelpath, index = False, startcol = 1, startrow = 1)
        infor = [self.executor, self.function, self.results.__len__(), self.success_count, self.failed_count, self.skip_count]
        lable = ["Người thực hiện", "Chức năng", "Số test case thực hiện", "Số test case thành công", "Số test case thất bại", "Số test case chưa test"]
        df = pd.DataFrame(infor, lable)
        workbook = openpyxl.load_workbook(self.excelpath)
        worksheet = workbook.active
        i = 2
        # Write DataFrame values to Excel
        for index, row in df.iterrows():
            worksheet['L' + f'{i}'] = index
            worksheet['M' + f'{i}'] = row[0]
            i = i + 1
        img = Image(self.logo)
        img.anchor = "L9"
        img.width = 240 
        img.height = 240

        worksheet.add_image(img)
        workbook.save(self.excelpath)


    def build_result(self, report, status, message):

        result = OrderedDict()
        names = mangle_test_address(report.nodeid)

        result['Tên testsuite'] = names[-2]
        result['Tên chức năng kiểm thử'] = names[-1]
        if report.test_doc is None:
          result['Mô tả'] = report.test_doc
        else:
          result['Mô tả'] = report.test_doc.strip()

        result['Kết quả']    = status
        result['Thời gian thực thi']  = getattr(report, 'duration', 0.0)
        result['Thời gian thực hiện'] = datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
        result['Thông báo']   = message
        result['Tên file'] = report.location[0]
        # result['Đánh dấu']   = report.test_marker
        self.append(result)


    def append_pass(self, report):
        status = "THÔNG QUA"
        message = None
        self.success_count += 1
        self.build_result(report, status, message)


    def append_failure(self, report):

        if hasattr(report, "wasxfail"):
            status = "XPASSED"
            message = "xfail-marked test passes Reason: %s " % report.wasxfail

        else:
            if hasattr(report.longrepr, "reprcrash"):
                message = report.longrepr.reprcrash.message
            elif isinstance(report.longrepr, (unicode, str)):
                message = report.longrepr
            else:
                message = str(report.longrepr)

            status = "THẤT BẠI"
            self.failed_count += 1
        self.build_result(report, status, message)


    def append_error(self, report):

        message = report.longrepr
        status = "LỖI"
        self.failed_count += 1
        self.build_result(report, status, message)


    def append_skipped(self, report):

        if hasattr(report, "wasxfail"):
            status = "XFAILED"
            message = "expected test failure Reason: %s " % report.wasxfail

        else:
            status = "SKIPPED"
            _, _, message = report.longrepr
            if message.startswith("Skipped: "):
                message = message[9:]
        self.skip_count += 1
        self.build_result(report, status, message)


    def build_tests(self, item):

        result = OrderedDict()
        names = mangle_test_address(item.nodeid)

        result['Tên testsuite'] = names[-2]
        result['Tên chức năng kiểm thử'] = names[-1]
        if item.obj.__doc__ is None:
          result['Mô tả'] = item.obj.__doc__
        else:
          result['Mô tả'] = item.obj.__doc__.strip()
        result['Tên file'] = item.location[0]
        test_marker = []
        test_message = []
        for k, v in item.keywords.items():
            if isinstance(v, list):
                for x in v:
                    if isinstance(x, Mark):
                        if x.name != 'usefixtures':
                            test_marker.append(x.name)
                        if x.kwargs:
                            test_message.append(x.kwargs.get('reason'))

        # test_markers = ', '.join(test_marker)
        # result['Đánh dấu'] = test_markers

        test_messages = ', '.join(test_message)
        result['Thông báo'] = test_messages
        self.append(result)


    def append_tests(self, item):

        self.build_tests(item)


    @pytest.hookimpl(trylast=True)
    def pytest_collection_modifyitems(self, session, config, items):
        """ called after collection has been performed, may filter or re-order
        the items in-place."""
        if session.config.option.collectonly:
            for item in items:
                self.append_tests(item)


    @pytest.hookimpl(hookwrapper=True)
    def pytest_runtest_makereport(self, item, call):

        outcome = yield

        report = outcome.get_result()
        report.test_doc = item.obj.__doc__
        test_marker = []
        for k, v in item.keywords.items():
            if isinstance(v,list):
                for x in v:
                    if isinstance(x,Mark):
                        test_marker.append(x.name)
        report.test_marker = ', '.join(test_marker)


    def pytest_runtest_logreport(self, report):

        if report.passed:
            if report.when == "call":  # ignore setup/teardown
                self.append_pass(report)

        elif report.failed:
            if report.when == "call":
                self.append_failure(report)

            else:
                self.append_error(report)

        elif report.skipped:
            self.append_skipped(report)


    def pytest_sessionfinish(self, session):
        if not hasattr(session.config, 'slaveinput'):
            if self.results:
                fieldnames = list(self.results[0])
                self.create_sheet(fieldnames)
                self.update_worksheet()
                self.save_excel()


    def pytest_terminal_summary(self, terminalreporter):
        if self.results:
            terminalreporter.write_sep("-", "excel report: %s" % (self.excelpath))